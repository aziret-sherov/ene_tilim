"""
Convert /Desktop/sozduk_words_full.xlsx → scripts/words.json
Row 3 = headers, rows 4+ = data.
Columns: word_kg, word_ru, word_en, category, example_kg, example_ru, example_en
"""
import json
import sys
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"], check=True)
    import openpyxl

EXCEL_PATH = "/Users/axelmacmini/Desktop/sozduk_words_full.xlsx"
OUT_PATH = "scripts/words.json"

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

# Detect format: if row 1 col 1 is "word_kg" headers start at row 1 (data row 2),
# otherwise old format with headers at row 3 (data row 4).
first_cell = str(ws.cell(1, 1).value or '').strip()
data_start = 2 if first_cell == 'word_kg' else 4

rows = []
skipped = 0
for row in ws.iter_rows(min_row=data_start, values_only=True):
    word_kg, word_ru, word_en, category, example_kg, example_ru, example_en = (row + (None,) * 7)[:7]
    if not word_kg or not str(word_kg).strip():
        skipped += 1
        continue
    entry = {
        "word_kg": str(word_kg).strip(),
        "word_ru": str(word_ru).strip() if word_ru else None,
        "word_en": str(word_en).strip() if word_en else None,
        "category": str(category).strip() if category else None,
        "example_kg": str(example_kg).strip() if example_kg else None,
        "example_ru": str(example_ru).strip() if example_ru else None,
        "example_en": str(example_en).strip() if example_en else None,
    }
    rows.append(entry)

with open(OUT_PATH, "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)

print(f"Converted {len(rows)} words → {OUT_PATH}  (skipped {skipped} empty rows)")
