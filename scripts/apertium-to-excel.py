"""
Parse apertium-eng-kir.eng-kir.dix and write an Excel file for review.
Columns: word_kg, word_ru, word_en, category, example_kg, example_ru, example_en

Usage:
    python scripts/apertium-to-excel.py
Output:
    ~/Desktop/apertium_words.xlsx
"""
import re
import sys
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"], check=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

INPUT = "/tmp/apertium-eng-kir.dix"
OUTPUT = "/Users/axelmacmini/Desktop/apertium_words.xlsx"

PAR_CATEGORY = {
    "par__n":      "noun",
    "par__num_pl": "numeral",
    "par__num_sg": "numeral",
}

entry_re = re.compile(r'<e[^>]*>\s*<p>\s*<l>([^<]+)</l>\s*<r>([^<]+)</r>\s*</p>\s*<par n="([^"]+)"', re.DOTALL)

# Group: word_kg -> {en: [list], category}
from collections import defaultdict
words: dict[str, dict] = {}

with open(INPUT, encoding="utf-8") as f:
    content = f.read()

for m in entry_re.finditer(content):
    word_en = m.group(1).strip()
    word_kg = m.group(2).strip()
    par     = m.group(3).strip()
    category = PAR_CATEGORY.get(par, "other")

    # Skip pure loanwords / transliterations that are identical in both langs
    if word_en.lower() == word_kg.lower():
        continue

    if word_kg not in words:
        words[word_kg] = {"en": [], "category": category}
    if word_en not in words[word_kg]["en"]:
        words[word_kg]["en"].append(word_en)

rows = sorted(words.items(), key=lambda x: x[0])

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "sozduk"

HEADERS = ["word_kg", "word_ru", "word_en", "category", "example_kg", "example_ru", "example_en"]
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF")

for col, h in enumerate(HEADERS, 1):
    cell = ws.cell(1, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for row_idx, (word_kg, data) in enumerate(rows, 2):
    word_en = ", ".join(data["en"])
    ws.cell(row_idx, 1, word_kg)
    ws.cell(row_idx, 2, "")           # word_ru — to fill
    ws.cell(row_idx, 3, word_en)
    ws.cell(row_idx, 4, data["category"])
    ws.cell(row_idx, 5, "")           # example_kg
    ws.cell(row_idx, 6, "")           # example_ru
    ws.cell(row_idx, 7, "")           # example_en

# Column widths
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 40
ws.column_dimensions["F"].width = 40
ws.column_dimensions["G"].width = 40

ws.freeze_panes = "A2"

wb.save(OUTPUT)
print(f"Done. {len(rows)} unique Kyrgyz words → {OUTPUT}")
